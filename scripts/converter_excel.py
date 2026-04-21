"""
Converte TABELA_ABRIL_2026.xlsx (planilha MAX) para data/produtos.json.
Uso: py scripts/converter_excel.py
Nunca inclui preços no JSON gerado.
"""

import json
import re
import unicodedata
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(__file__).parent.parent / "data" / "TABELA_ABRIL_2026.xlsx"
OUTPUT_PATH = Path(__file__).parent.parent / "data" / "produtos.json"
SHEET_NAME = "MAX"

HEADER_KEYWORDS = {"CDG", "PRODUTO", "LAR", "COM", "ESP", "PREÇO", "MIL/FRD",
                   "CAIXA", "VALOR", "QUANTIDADE", "UNIDADE", "TAMANHO"}

PRICE_COL_NAMES = {"PRE\ufffdO", "PREÇO", "VALOR", "PRE"}


def to_slug(text: str) -> str:
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def is_category_row(row: tuple) -> bool:
    """Linha de categoria: texto em col 0 ou 1, demais None ou palavras de cabeçalho."""
    # Coletar valores não-None ignorando palavras-chave de cabeçalho
    meaningful = []
    for v in row:
        if v is None:
            continue
        s = str(v).strip().upper()
        if s in HEADER_KEYWORDS or s == "MIL/FRD":
            continue
        meaningful.append(str(v).strip())

    if len(meaningful) != 1:
        return False
    text = meaningful[0]
    if not text or len(text) < 3:
        return False
    if text.upper() in HEADER_KEYWORDS:
        return False
    # Não pode ser só número
    if re.match(r"^\d+$", text):
        return False
    # Texto da categoria deve estar em col 0 ou 1
    if row[0] is not None:
        candidate = str(row[0]).strip()
    elif len(row) > 1 and row[1] is not None:
        candidate = str(row[1]).strip()
    else:
        return False
    return candidate.upper() not in HEADER_KEYWORDS


def is_header_row(row: tuple) -> bool:
    """Linha de cabeçalho de colunas (CDG, PRODUTO, etc.)"""
    vals = [str(v).strip().upper() for v in row if v is not None]
    return any(k in vals for k in ("CDG", "PRODUTO"))


def is_product_row(row: tuple) -> bool:
    """Linha de produto: tem código (col 0 ou 1) e nome de produto."""
    codigo = str(row[0]).strip() if row[0] is not None else ""
    nome = str(row[1]).strip() if row[1] is not None else ""
    if not nome or len(nome) < 2:
        return False
    if nome.upper() in HEADER_KEYWORDS:
        return False
    if is_category_row(row) or is_header_row(row):
        return False
    return True


def clean_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def parse_sheet(ws) -> list:
    categorias = []
    current_cat = None

    for row in ws.iter_rows(values_only=True):
        # Ignorar linhas totalmente vazias
        if not any(v is not None for v in row):
            continue

        if is_header_row(row):
            continue

        if is_category_row(row):
            # Texto da categoria pode estar em col 0 ou col 1
            nome_cat = clean_str(row[1] if row[0] is None else row[0])
            if not nome_cat:
                continue
            current_cat = {
                "slug": to_slug(nome_cat),
                "nome": nome_cat.title(),
                "produtos": []
            }
            categorias.append(current_cat)
            continue

        if current_cat is None:
            continue

        if is_product_row(row):
            codigo = clean_str(row[0])
            nome = clean_str(row[1])

            # Dimensões (largura x comprimento) quando disponíveis
            largura = clean_str(row[2])
            comprimento = clean_str(row[3])
            espessura = clean_str(row[4])

            dimensoes = ""
            if largura and comprimento:
                try:
                    l = int(float(largura))
                    c = int(float(comprimento))
                    dimensoes = f"{l} x {c} cm"
                    if espessura:
                        dimensoes += f" — {espessura} mm"
                except (ValueError, TypeError):
                    pass

            # Unidade/embalagem (col 3 ou 4 em categorias sem dimensão)
            unidade = ""
            if not dimensoes:
                for idx in (3, 4):
                    v = clean_str(row[idx]) if idx < len(row) else ""
                    if v and v.upper() not in HEADER_KEYWORDS:
                        unidade = v
                        break

            produto = {
                "codigo": codigo if codigo else None,
                "nome": nome,
                "slug_categoria": current_cat["slug"],
                "imagem": f"imagens/produtos/{current_cat['slug']}/{to_slug(nome)}.jpg"
            }
            if dimensoes:
                produto["dimensoes"] = dimensoes
            if unidade:
                produto["unidade"] = unidade

            # Remover chaves com valor None
            produto = {k: v for k, v in produto.items() if v is not None and v != ""}
            current_cat["produtos"].append(produto)

    # Remover categorias sem produtos
    categorias = [c for c in categorias if c["produtos"]]
    return categorias


def main():
    print(f"Lendo {EXCEL_PATH} …")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Planilha '{SHEET_NAME}' não encontrada. Disponíveis: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    categorias = parse_sheet(ws)
    wb.close()

    total_produtos = sum(len(c["produtos"]) for c in categorias)
    print(f"Extraídas {len(categorias)} categorias e {total_produtos} produtos.")

    output = {"categorias": categorias}
    OUTPUT_PATH.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"JSON salvo em {OUTPUT_PATH}")

    for cat in categorias:
        print(f"  {cat['nome']:40s} — {len(cat['produtos'])} produtos")


if __name__ == "__main__":
    main()
